from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from models import Factura, Paquete, Cliente, db
from datetime import datetime
import io
import openpyxl

facturas_bp = Blueprint('facturas', __name__, url_prefix='/facturas')

@facturas_bp.route('/')
@login_required
def index():
    estado = request.args.get('estado', '')
    query = Factura.query.options(joinedload(Factura.cliente), joinedload(Factura.paquetes)).join(Cliente)
    if estado:
        query = query.filter(Factura.estado == estado)
    facturas = query.order_by(Factura.fecha_emision.desc()).all()
    return render_template('facturas/index.html', facturas=facturas, estado=estado)

@facturas_bp.route('/nueva', methods=['GET', 'POST'])
@login_required
def nueva():
    clientes = Cliente.query.filter_by(activo=True).order_by(Cliente.nombre_completo).all()
    cliente_id = request.args.get('cliente_id')

    if request.method == 'POST':
        cliente_id = int(request.form.get('cliente_id'))
        paquete_ids = request.form.getlist('paquete_ids')

        if not paquete_ids:
            flash('Debes seleccionar al menos un paquete.', 'error')
            return redirect(request.url)

        factura = Factura(
            numero=Factura.generar_numero(),
            cliente_id=cliente_id,
            notas=request.form.get('notas', ''),
            creado_por=current_user.id,
            estado='borrador'
        )
        db.session.add(factura)
        db.session.flush()

        facturas_afectadas = set()
        
        # Optimization: Fetch all packages at once (prevents N+1 queries)
        if paquete_ids:
            paquetes_a_actualizar = Paquete.query.filter(Paquete.id.in_([int(p) for p in paquete_ids])).all()
            for paquete in paquetes_a_actualizar:
                if paquete.factura_id and paquete.factura_id != factura.id:
                    facturas_afectadas.add(paquete.factura_id)
                paquete.factura_id = factura.id

        db.session.flush()

        if facturas_afectadas:
            # Optimization: Fetch all affected facturas at once
            facturas_antiguas = Factura.query.filter(Factura.id.in_(list(facturas_afectadas))).all()
            for f_ant in facturas_antiguas:
                if Paquete.query.filter_by(factura_id=f_ant.id).count() == 0:
                    from models import Pago
                    Pago.query.filter_by(factura_id=f_ant.id).delete()
                    db.session.delete(f_ant)
                else:
                    # Optimization: manually update total to prevent mid-loop db.session.commit() from f_ant.actualizar_total()
                    f_ant.total = f_ant.calcular_total()

        # Update current factura without intermediate commit
        factura.total = factura.calcular_total()
        db.session.commit()
        
        from models import registrar_actividad
        registrar_actividad(current_user.id, 'Creó Factura', f'Factura {factura.numero} por ${factura.total:.2f}')
        
        flash(f'Factura {factura.numero} creada.', 'success')
        return redirect(url_for('facturas.detalle', id=factura.id))

    return render_template('facturas/nueva.html', clientes=clientes, cliente_id=cliente_id)

@facturas_bp.route('/detalle/<int:id>')
@login_required
def detalle(id):
    factura = Factura.query.get_or_404(id)
    # WhatsApp message
    tel = (factura.cliente.telefono or '').replace('-', '').replace(' ', '')
    msg = (f"Hola {factura.cliente.nombre_completo}, adjunto su factura {factura.numero} "
           f"por un total de ${factura.total:.2f}. Gracias por preferirnos.")
    import urllib.parse
    wa_url = f"https://wa.me/505{tel}?text={urllib.parse.quote(msg)}" if tel else f"https://wa.me/?text={urllib.parse.quote(msg)}"
    return render_template('facturas/detalle.html', factura=factura, whatsapp_url=wa_url)

@facturas_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    if current_user.rol != 'admin':
        flash('No tienes permisos para editar facturas.', 'error')
        return redirect(url_for('facturas.index'))
        
    factura = Factura.query.get_or_404(id)

    if request.method == 'POST':
        factura.notas = request.form.get('notas', '')
        # Quitar paquetes actuales
        for p in factura.paquetes:
            p.factura_id = None
            
        # Asignar nuevos
        paquete_ids = request.form.getlist('paquete_ids')
        facturas_afectadas = set()
        
        if paquete_ids:
            paquetes_a_actualizar = Paquete.query.filter(Paquete.id.in_([int(p) for p in paquete_ids])).all()
            for paquete in paquetes_a_actualizar:
                if paquete.factura_id and paquete.factura_id != factura.id:
                    facturas_afectadas.add(paquete.factura_id)
                paquete.factura_id = factura.id
                
        db.session.flush()
        
        if facturas_afectadas:
            facturas_antiguas = Factura.query.filter(Factura.id.in_(list(facturas_afectadas))).all()
            for f_ant in facturas_antiguas:
                if Paquete.query.filter_by(factura_id=f_ant.id).count() == 0:
                    from models import Pago
                    Pago.query.filter_by(factura_id=f_ant.id).delete()
                    db.session.delete(f_ant)
                else:
                    f_ant.total = f_ant.calcular_total()

        factura.total = factura.calcular_total()
        db.session.commit()
        flash('Factura actualizada.', 'success')
        return redirect(url_for('facturas.detalle', id=id))

    cliente = factura.cliente
    from sqlalchemy import or_
    paquetes_sin_facturar = Paquete.query.outerjoin(Factura).filter(
        Paquete.cliente_id == cliente.id,
        or_(
            Paquete.factura_id == None,
            db.and_(Factura.estado != 'pagada', Factura.id != factura.id)
        )
    ).all()
    return render_template('facturas/editar.html', factura=factura, paquetes_sin_facturar=paquetes_sin_facturar)

@facturas_bp.route('/finalizar/<int:id>', methods=['POST'])
@login_required
def finalizar(id):
    factura = Factura.query.get_or_404(id)
    factura.estado = 'finalizada'
    db.session.commit()
    
    from models import registrar_actividad
    registrar_actividad(current_user.id, 'Finalizó Factura', f'Factura {factura.numero} marcada como finalizada')
    
    flash('Factura finalizada.', 'success')
    return redirect(url_for('facturas.detalle', id=id))

@facturas_bp.route('/pagar/<int:id>', methods=['POST'])
@login_required
def pagar(id):
    from models import Pago
    factura = Factura.query.get_or_404(id)
    if factura.estado == 'finalizada':
        metodo_pago = request.form.get('metodo_pago', 'Efectivo')
        referencia = request.form.get('referencia', '')
        
        pago = Pago(
            factura_id=factura.id,
            monto=factura.total,
            metodo_pago=metodo_pago,
            referencia=referencia,
            registrado_por=current_user.id
        )
        db.session.add(pago)
        factura.estado = 'pagada'
        db.session.commit()
        
        from models import registrar_actividad
        registrar_actividad(current_user.id, 'Registró Pago', f'Pago de ${factura.total:.2f} con {metodo_pago} en factura {factura.numero}')
        
        flash(f'Factura marcada como pagada con {metodo_pago}.', 'success')
    else:
        flash('Solo se pueden pagar facturas finalizadas.', 'warning')
    return redirect(url_for('facturas.detalle', id=id))

@facturas_bp.route('/pdf/<int:id>')
@login_required
def generar_pdf(id):
    factura = Factura.query.get_or_404(id)
    pdf_bytes = generar_pdf_factura(factura)
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=factura-{factura.numero}.pdf'
    return response

@facturas_bp.route('/paquetes-cliente/<int:cliente_id>')
@login_required
def paquetes_cliente(cliente_id):
    from sqlalchemy import or_
    paquetes = Paquete.query.outerjoin(Factura).filter(
        Paquete.cliente_id == cliente_id,
        or_(
            Paquete.factura_id == None,
            Factura.estado != 'pagada'
        )
    ).all()
    return jsonify([{
        'id': p.id, 'nombre': p.nombre, 'peso': p.peso,
        'tipo_envio': p.tipo_envio, 'costo': p.costo,
        'factura_previa': p.factura.numero if p.factura else None,
        'numero_seguimiento': p.numero_seguimiento,
        'tracking_number': p.tracking_number,
        'estado_rastreo': p.estado_rastreo
    } for p in paquetes])

@facturas_bp.route('/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar(id):
    if current_user.rol != 'admin':
        flash('No tienes permisos para eliminar facturas.', 'error')
        return redirect(url_for('facturas.index'))
        
    factura = Factura.query.get_or_404(id)
        
    for p in factura.paquetes:
        p.factura_id = None
        
    from models import Pago
    pagos = Pago.query.filter_by(factura_id=factura.id).all()
    for pago in pagos:
        db.session.delete(pago)
        
    numero = factura.numero
    db.session.delete(factura)
    db.session.commit()
    
    from models import registrar_actividad
    registrar_actividad(current_user.id, 'Eliminó Factura', f'Factura {numero} eliminada')
    
    flash('Factura eliminada correctamente.', 'success')
    return redirect(url_for('facturas.index'))

def generar_pdf_factura(factura):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    import os

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch,
                            title=f"Factura {factura.numero}",
                            author="BJJ SYSTEM")

    styles = getSampleStyleSheet()
    COLOR_PRIMARIO = colors.HexColor('#3d5ba0')
    COLOR_ACENTO = colors.HexColor('#3d5ba0')

    story = []

    # Encabezado con Logo
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'logoazul.PNG')
    
    if os.path.exists(logo_path):
        header_left = Image(logo_path, width=3.5*inch, height=1.2*inch, kind='proportional')
        header_left.hAlign = 'LEFT'
    else:
        header_left = Paragraph('<font size=22 color="#3d5ba0"><b>ENVÍOS BJJ</b></font>', styles['Normal'])

    company_info_text = (
        '<font size=9 color="#333333">'
        '<b>RUC:</b> 001-220905-1038A<br/>'
        '<b>Dirección:</b> Villa flor Norte, Iglesia Catolica, 3C N, 10vrs E. Casa 1031'
        '</font>'
    )

    header_data = [
        [header_left,
         Paragraph(f'<font size=10 color="#666666">FACTURA<br/><font size=16 color="#3d5ba0"><b>{factura.numero}</b></font></font>', styles['Normal'])]
    ]
    header_table = Table(header_data, colWidths=[4*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.1*inch))
    story.append(Paragraph(company_info_text, styles['Normal']))
    story.append(Spacer(1, 0.2*inch))

    # Línea separadora
    line = Table([['']], colWidths=[7*inch], rowHeights=[3])
    line.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), COLOR_ACENTO)]))
    story.append(line)
    story.append(Spacer(1, 0.2*inch))

    # Info cliente y factura
    info_data = [
        [Paragraph('<b>FACTURAR A:</b>', styles['Normal']),
         Paragraph('<b>DETALLES:</b>', styles['Normal'])],
        [Paragraph(factura.cliente.nombre_completo, styles['Normal']),
         Paragraph(f'Fecha: {factura.fecha_emision.strftime("%d/%m/%Y")}', styles['Normal'])],
        [Paragraph(f'Cédula: {factura.cliente.cedula}', styles['Normal']),
         Paragraph(f'Estado: {factura.estado.upper()}', styles['Normal'])],
        [Paragraph(f'Tel: {factura.cliente.telefono or "-"}', styles['Normal']), Paragraph('', styles['Normal'])],
        [Paragraph(f'Email: {factura.cliente.email or "-"}', styles['Normal']), Paragraph('', styles['Normal'])],
    ]
    info_table = Table(info_data, colWidths=[3.5*inch, 3.5*inch])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))

    # Tabla de paquetes
    table_data = [['#', 'Paquete', 'Guía de Rastreo', 'Tipo', 'Peso (lb)', 'Costo']]
    
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=9, wordWrap='CJK')
    
    for i, p in enumerate(factura.paquetes, 1):
        tipo = 'Aéreo' if p.tipo_envio == 'aereo' else 'Marítimo'
        nombre_p = Paragraph(p.nombre, cell_style)
        rastreo_p = Paragraph(p.numero_seguimiento or '—', cell_style)
        table_data.append([
            str(i), nombre_p, rastreo_p, tipo,
            f'{p.peso:.2f}', f'${p.costo:.2f}'
        ])

    col_widths = [0.4*inch, 1.8*inch, 1.8*inch, 1*inch, 0.9*inch, 0.9*inch]
    paquetes_table = Table(table_data, colWidths=col_widths)
    paquetes_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (4,0), (5,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(paquetes_table)
    story.append(Spacer(1, 0.2*inch))

    # Total
    total_cordobas = factura.total * 37
    total_data = [
        ['', '', '', '', 'TOTAL $:', f'${factura.total:.2f}'],
        ['', '', '', '', 'TOTAL C$:', f'C${total_cordobas:.2f}']
    ]
    total_table = Table(total_data, colWidths=col_widths)
    total_table.setStyle(TableStyle([
        ('FONTNAME', (4,0), (5,1), 'Helvetica-Bold'),
        ('FONTSIZE', (4,0), (5,0), 11),
        ('FONTSIZE', (4,1), (5,1), 10),
        ('ALIGN', (4,0), (4,1), 'LEFT'),
        ('ALIGN', (5,0), (5,1), 'RIGHT'),
        ('BACKGROUND', (4,0), (5,1), COLOR_ACENTO),
        ('TEXTCOLOR', (4,0), (5,1), colors.white),
        ('TOPPADDING', (4,0), (5,1), 6),
        ('BOTTOMPADDING', (4,0), (5,1), 6),
        ('LINEBELOW', (4,0), (5,0), 0.5, colors.white),
    ]))
    story.append(total_table)

    if factura.notas:
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph(f'<b>Notas:</b> {factura.notas}', styles['Normal']))

    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph('<font size=12 color="#000000"><b>Gracias por confiar su carga con nosotros.</b></font>', styles['Normal']))

    doc.build(story)
    return buffer.getvalue()

@facturas_bp.route('/exportar')
@login_required
def exportar():
    if current_user.rol != 'admin':
        flash('No tienes permisos para exportar datos.', 'error')
        return redirect(url_for('facturas.index'))
        
    estado = request.args.get('estado', '')
    filtro_fecha = request.args.get('filtro_fecha', '')
    dia = request.args.get('dia', '')
    mes = request.args.get('mes', '')
    semana = request.args.get('semana', '')

    query = Factura.query.options(joinedload(Factura.cliente), joinedload(Factura.paquetes)).join(Cliente)
    if estado:
        query = query.filter(Factura.estado == estado)

    if filtro_fecha == 'dia' and dia:
        from datetime import datetime, time
        target_date = datetime.strptime(dia, '%Y-%m-%d').date()
        start_date = datetime.combine(target_date, time.min)
        end_date = datetime.combine(target_date, time.max)
        query = query.filter(Factura.fecha_emision >= start_date, Factura.fecha_emision <= end_date)
    elif filtro_fecha == 'mes' and mes:
        year, month = map(int, mes.split('-'))
        import calendar
        from datetime import datetime
        _, last_day = calendar.monthrange(year, month)
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, last_day, 23, 59, 59)
        query = query.filter(Factura.fecha_emision >= start_date, Factura.fecha_emision <= end_date)
    elif filtro_fecha == 'semana' and semana:
        from datetime import datetime, timedelta
        year_str, week_str = semana.split('-W')
        start_date = datetime.strptime(semana + '-1', '%G-W%V-%u')
        end_date = start_date + timedelta(days=7)
        query = query.filter(Factura.fecha_emision >= start_date, Factura.fecha_emision < end_date)

    facturas = query.order_by(Factura.fecha_emision.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    headers = ['ID', 'Número Factura', 'Cliente', 'Cédula', 'Fecha Emisión', 'Estado', 'Total Facturado ($)', 'Total Facturado (C$)', 'Costo Agencia ($)', 'Ganancia ($)', 'Cant. Paquetes', 'Notas']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='3D5BA0', end_color='3D5BA0', fill_type='solid')

    for f in facturas:
        costo_agencia = sum((p.peso * 5.0) if p.tipo_envio == 'aereo' else (p.peso * 1.6) for p in f.paquetes)
        ganancia = f.total - costo_agencia
        
        ws.append([
            f.id,
            f.numero,
            f.cliente.nombre_completo,
            f.cliente.cedula,
            f.fecha_emision.strftime('%Y-%m-%d %H:%M') if f.fecha_emision else '',
            f.estado.upper(),
            f.total,
            f.total * 37,
            costo_agencia,
            ganancia,
            len(f.paquetes),
            f.notas or ''
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from models import get_local_now
    filename = f"Facturas_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
