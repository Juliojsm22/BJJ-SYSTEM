from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from models import Cliente, db
import io
import openpyxl
import re

clientes_bp = Blueprint('clientes', __name__, url_prefix='/clientes')

@clientes_bp.route('/')
@login_required
def index():
    q = request.args.get('q', '')
    query = Cliente.query.options(joinedload(Cliente.paquetes)).filter_by(activo=True)
    if q:
        clientes = query.filter(
            (Cliente.nombre_completo.ilike(f'%{q}%')) |
            (Cliente.cedula.ilike(f'%{q}%')) |
            (Cliente.email.ilike(f'%{q}%'))
        ).order_by(Cliente.nombre_completo).all()
    else:
        clientes = query.order_by(Cliente.nombre_completo).all()
    return render_template('clientes/index.html', clientes=clientes, q=q)

@clientes_bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo():
    if request.method == 'POST':
        cedula = request.form.get('cedula').strip().upper()
        if not re.match(r'^\d{3}-\d{6}-\d{4}[A-Z]$', cedula):
            flash('La cédula debe tener el formato de Nicaragua: 000-000000-0000A', 'error')
            return render_template('clientes/form.html', cliente=None)
            
        if Cliente.query.filter_by(cedula=cedula).first():
            flash('Ya existe un cliente con esa cédula.', 'error')
            return render_template('clientes/form.html', cliente=None)
        
        telefono = request.form.get('telefono', '').strip()
        if telefono and (len(telefono) != 8 or not telefono.isdigit()):
            flash('El número telefónico debe contener exactamente 8 dígitos numéricos.', 'error')
            return render_template('clientes/form.html', cliente=None)
            
        cliente = Cliente(
            nombre_completo=request.form.get('nombre_completo').strip(),
            cedula=cedula,
            telefono=telefono,
            email=request.form.get('email').strip()
        )
        db.session.add(cliente)
        db.session.flush() # Para obtener el ID del cliente
        
        tarifa_aereo = request.form.get('tarifa_aereo')
        tarifa_maritimo = request.form.get('tarifa_maritimo')
        
        if tarifa_aereo or tarifa_maritimo:
            from models import TarifaEspecialCliente
            tarifa_esp = TarifaEspecialCliente(
                cliente_id=cliente.id,
                aereo=float(tarifa_aereo) if tarifa_aereo else None,
                maritimo=float(tarifa_maritimo) if tarifa_maritimo else None
            )
            db.session.add(tarifa_esp)
            
        db.session.commit()
        
        from models import registrar_actividad
        registrar_actividad(current_user.id, 'Registró Cliente', f'Cliente {cliente.nombre_completo} (Cédula: {cliente.cedula})')
        
        flash('Cliente registrado exitosamente.', 'success')
        return redirect(url_for('clientes.index'))
    
    return render_template('clientes/form.html', cliente=None)

@clientes_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    cliente = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        cedula = request.form.get('cedula').strip().upper()
        if not re.match(r'^\d{3}-\d{6}-\d{4}[A-Z]$', cedula):
            flash('La cédula debe tener el formato de Nicaragua: 000-000000-0000A', 'error')
            return render_template('clientes/form.html', cliente=cliente)
            
        existente = Cliente.query.filter_by(cedula=cedula).first()
        if existente and existente.id != id:
            flash('Ya existe un cliente con esa cédula.', 'error')
            return render_template('clientes/form.html', cliente=cliente)
        
        telefono = request.form.get('telefono', '').strip()
        if telefono and (len(telefono) != 8 or not telefono.isdigit()):
            flash('El número telefónico debe contener exactamente 8 dígitos numéricos.', 'error')
            return render_template('clientes/form.html', cliente=cliente)
            
        cliente.nombre_completo = request.form.get('nombre_completo').strip()
        cliente.cedula = cedula
        cliente.telefono = telefono
        cliente.email = request.form.get('email').strip()
        
        tarifa_aereo = request.form.get('tarifa_aereo')
        tarifa_maritimo = request.form.get('tarifa_maritimo')
        
        from models import TarifaEspecialCliente, Paquete, Tarifa
        if tarifa_aereo or tarifa_maritimo:
            if not cliente.tarifa_especial:
                cliente.tarifa_especial = TarifaEspecialCliente(cliente_id=cliente.id)
            cliente.tarifa_especial.aereo = float(tarifa_aereo) if tarifa_aereo else None
            cliente.tarifa_especial.maritimo = float(tarifa_maritimo) if tarifa_maritimo else None
        else:
            if cliente.tarifa_especial:
                db.session.delete(cliente.tarifa_especial)
                
        # Recalcular costos de paquetes pendientes
        t_aereo = Tarifa.query.filter_by(nombre='aereo').first()
        t_maritimo = Tarifa.query.filter_by(nombre='maritimo').first()
        precio_aereo_base = t_aereo.precio_por_libra if t_aereo else 6.50
        precio_maritimo_base = t_maritimo.precio_por_libra if t_maritimo else 2.50
        
        paquetes_pendientes = Paquete.query.filter_by(cliente_id=cliente.id, factura_id=None).all()
        for p in paquetes_pendientes:
            tarifa_p = None
            if cliente.tarifa_especial:
                if p.tipo_envio == 'aereo' and cliente.tarifa_especial.aereo is not None:
                    tarifa_p = cliente.tarifa_especial.aereo
                elif p.tipo_envio == 'maritimo' and cliente.tarifa_especial.maritimo is not None:
                    tarifa_p = cliente.tarifa_especial.maritimo
            
            if tarifa_p is None:
                tarifa_p = precio_aereo_base if p.tipo_envio == 'aereo' else precio_maritimo_base
                
            p.costo = round(p.peso * tarifa_p, 2)

        db.session.commit()
        
        from models import registrar_actividad
        registrar_actividad(current_user.id, 'Editó Cliente', f'Actualizó perfil de {cliente.nombre_completo}')
        
        flash('Cliente actualizado correctamente.', 'success')
        return redirect(url_for('clientes.index'))
    
    return render_template('clientes/form.html', cliente=cliente)

@clientes_bp.route('/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar(id):
    if current_user.rol != 'admin':
        flash('No tienes permisos para esto.', 'error')
        return redirect(url_for('clientes.index'))
    cliente = Cliente.query.get_or_404(id)
    cliente.activo = False
    db.session.commit()
    
    from models import registrar_actividad
    registrar_actividad(current_user.id, 'Eliminó Cliente', f'Desactivó a {cliente.nombre_completo}')
    
    flash('Cliente eliminado.', 'success')
    return redirect(url_for('clientes.index'))

@clientes_bp.route('/buscar-json')
@login_required
def buscar_json():
    q = request.args.get('q', '')
    clientes = Cliente.query.filter(
        (Cliente.nombre_completo.ilike(f'%{q}%')) | (Cliente.cedula.ilike(f'%{q}%'))
    ).filter_by(activo=True).limit(10).all()
    return jsonify([{'id': c.id, 'nombre': c.nombre_completo, 'cedula': c.cedula} for c in clientes])

@clientes_bp.route('/detalle/<int:id>')
@login_required
def detalle(id):
    cliente = Cliente.query.get_or_404(id)
    return render_template('clientes/detalle.html', cliente=cliente)

@clientes_bp.route('/exportar')
@login_required
def exportar():
    if current_user.rol != 'admin':
        flash('No tienes permisos para exportar datos.', 'error')
        return redirect(url_for('clientes.index'))
        
    q = request.args.get('q', '')
    filtro_fecha = request.args.get('filtro_fecha', '')
    mes = request.args.get('mes', '')
    semana = request.args.get('semana', '')

    query = Cliente.query.options(joinedload(Cliente.paquetes)).filter_by(activo=True)
    if q:
        query = query.filter(
            (Cliente.nombre_completo.ilike(f'%{q}%')) |
            (Cliente.cedula.ilike(f'%{q}%')) |
            (Cliente.email.ilike(f'%{q}%'))
        )
        
    if filtro_fecha == 'mes' and mes:
        year, month = map(int, mes.split('-'))
        import calendar
        from datetime import datetime
        _, last_day = calendar.monthrange(year, month)
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, last_day, 23, 59, 59)
        query = query.filter(Cliente.creado_en >= start_date, Cliente.creado_en <= end_date)
    elif filtro_fecha == 'semana' and semana:
        from datetime import datetime, timedelta
        year_str, week_str = semana.split('-W')
        start_date = datetime.strptime(semana + '-1', '%G-W%V-%u')
        end_date = start_date + timedelta(days=7)
        query = query.filter(Cliente.creado_en >= start_date, Cliente.creado_en < end_date)

    clientes = query.order_by(Cliente.nombre_completo).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Encabezados
    headers = ['ID', 'Nombre Completo', 'Cédula', 'Teléfono', 'Email', 'Paquetes Sin Facturar', 'Total Libras Histórico', 'Fecha Registro']
    ws.append(headers)

    # Formato de encabezados
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='3D5BA0', end_color='3D5BA0', fill_type='solid')

    for c in clientes:
        ws.append([
            c.id,
            c.nombre_completo,
            c.cedula,
            c.telefono or '',
            c.email or '',
            len(c.paquetes_sin_facturar),
            c.total_libras,
            c.creado_en.strftime('%Y-%m-%d %H:%M') if c.creado_en else ''
        ])

    # Autoajustar columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from models import get_local_now
    filename = f"Clientes_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
