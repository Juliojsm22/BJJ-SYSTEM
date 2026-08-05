from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, make_response
from markupsafe import Markup
from flask_login import login_required, current_user
from sqlalchemy.orm import joinedload
from models import Paquete, Cliente, db, HistorialRastreo, Tarifa
import io
import openpyxl
import os
import re

paquetes_bp = Blueprint('paquetes', __name__, url_prefix='/paquetes')

@paquetes_bp.route('/')
@login_required
def index():
    q = request.args.get('q', '')
    tipo = request.args.get('tipo', '')
    estado = request.args.get('estado', '')

    query = Paquete.query.options(joinedload(Paquete.cliente), joinedload(Paquete.factura)).join(Cliente).filter(Cliente.activo == True)
    if q:
        q_clean = re.sub(r'[^a-zA-Z0-9]', '', q).upper()
        base_wr = re.search(r'WR\d+', q_clean, flags=re.I)
        base_wr_str = base_wr.group(0).upper() if base_wr else q
        
        query = query.filter(
            (Paquete.nombre.ilike(f'%{q}%')) |
            (Cliente.nombre_completo.ilike(f'%{q}%')) |
            (Paquete.warehouse.ilike(f'%{q}%')) |
            (Paquete.warehouse.ilike(f'%{base_wr_str}%')) |
            (Paquete.numero_seguimiento.ilike(f'%{q}%')) |
            (Paquete.tracking_number.ilike(f'%{q}%'))
        )
    if tipo:
        query = query.filter(Paquete.tipo_envio == tipo)
    if estado == 'sin_facturar':
        query = query.filter(Paquete.factura_id == None)
    elif estado == 'facturado':
        query = query.filter(Paquete.factura_id != None)

    paquetes = query.order_by(Paquete.registrado_en.desc()).all()
    return render_template('paquetes/index.html', paquetes=paquetes, q=q, tipo=tipo, estado=estado)

@paquetes_bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo():
    clientes = Cliente.query.filter_by(activo=True).order_by(Cliente.nombre_completo).all()
    cliente_id = request.args.get('cliente_id')

    if request.method == 'POST':
        cliente_id_form = int(request.form.get('cliente_id'))
        cliente = Cliente.query.get(cliente_id_form)
        
        nombres = request.form.getlist('nombre[]')
        descripciones = request.form.getlist('descripcion[]')
        pesos = request.form.getlist('peso[]')
        tipos_envio = request.form.getlist('tipo_envio[]')
        numeros_seguimiento = request.form.getlist('numero_seguimiento[]')
        warehouses = request.form.getlist('warehouse[]')
        estados_rastreo = request.form.getlist('estado_rastreo[]')
        
        # Validar números de seguimiento duplicados antes de guardar
        numeros_vistos = set()
        for num in numeros_seguimiento:
            num_strip = num.strip()
            if num_strip:
                if num_strip in numeros_vistos:
                    flash(f'El número de seguimiento "{num_strip}" está duplicado en el formulario.', 'error')
                    return redirect(request.url)
                numeros_vistos.add(num_strip)
                
                existente = Paquete.query.filter_by(numero_seguimiento=num_strip).first()
                if existente:
                    flash(f'El número de seguimiento "{num_strip}" ya está registrado en el paquete {existente.tracking_number}.', 'error')
                    return redirect(request.url)
                    
        paquetes_creados = []
        costo_total = 0
        
        for i in range(len(nombres)):
            peso = int(pesos[i] if pesos[i] else 0)
            tipo_envio = tipos_envio[i]
            numero_seg = numeros_seguimiento[i].strip()
            
            tarifa = None
            if cliente and cliente.tarifa_especial:
                if tipo_envio == 'aereo' and cliente.tarifa_especial.aereo is not None:
                    tarifa = cliente.tarifa_especial.aereo
                elif tipo_envio == 'maritimo' and cliente.tarifa_especial.maritimo is not None:
                    tarifa = cliente.tarifa_especial.maritimo
                    
            if tarifa is None:
                tarifa_db = Tarifa.query.filter_by(nombre=tipo_envio).first()
                tarifa = tarifa_db.precio_por_libra if tarifa_db else (6.50 if tipo_envio == 'aereo' else 2.50)
                
            paquete = Paquete(
                nombre=nombres[i].strip(),
                descripcion=descripciones[i].strip() if i < len(descripciones) else '',
                peso=peso,
                tipo_envio=tipo_envio,
                cliente_id=cliente_id_form,
                numero_seguimiento=numero_seg,
                warehouse=warehouses[i].strip() if i < len(warehouses) and warehouses[i].strip() else None,
                estado_rastreo=estados_rastreo[i] if i < len(estados_rastreo) else 'bodega_miami',
                registrado_por=current_user.id
            )
            paquete.save()
            costo_total += paquete.costo
            paquetes_creados.append(paquete)
            
            historial_inicial = HistorialRastreo(
                paquete_id=paquete.id,
                estado=paquete.estado_rastreo,
                ubicacion='Miami',
                comentarios='Paquete registrado en el sistema',
                creado_por=current_user.id
            )
            db.session.add(historial_inicial)
            
        db.session.commit()
        
        if paquetes_creados:
            from models import registrar_actividad
            registrar_actividad(current_user.id, 'Registró Paquetes', f'Registró {len(paquetes_creados)} paquete(s) para {cliente.nombre_completo}')
            
            if paquetes_creados and cliente.telefono:
                import urllib.parse
                mensaje = f"👋 *Hola {cliente.nombre_completo}*,\n\nTe informamos sobre el registro de tu(s) paquete(s) en *BJJ SYSTEM*.\n\n📦 *Detalles:*\n"
                for p in paquetes_creados:
                    if p.estado_rastreo == 'bodega_miami':
                        estado_str = "🏢🇺🇸 Bodega Miami"
                    elif p.estado_rastreo == 'en_transito':
                        estado_str = "🚢/✈️ En tránsito"
                    elif p.estado_rastreo == 'listo_para_retirar':
                        estado_str = "✅ Listo para retirar"
                    else:
                        estado_str = p.estado_rastreo.replace('_', ' ').title()
                    
                    track_url = url_for('rastreo.index', codigo=p.tracking_number, _external=True)
                    mensaje += f"🔸 *{p.nombre}* ({estado_str})\nTracking: {p.numero_seguimiento or p.tracking_number}\nPeso: {p.peso} lb\n🔗 Rastreo: {track_url}\n\n"
                mensaje += "¡Gracias por preferir *BJJ SYSTEM*! 🚀"
                
                telefono_limpio = ''.join(filter(str.isdigit, str(cliente.telefono)))
                if telefono_limpio:
                    wa_url = f"https://api.whatsapp.com/send?phone={telefono_limpio}&text={urllib.parse.quote(mensaje)}"
                    flash(wa_url, 'whatsapp')
            
            accion = request.form.get('accion')
            if accion == 'facturar':
                from models import Factura
                factura = Factura(
                    numero=Factura.generar_numero(),
                    cliente_id=cliente.id,
                    notas='',
                    creado_por=current_user.id,
                    estado='borrador'
                )
                db.session.add(factura)
                db.session.flush()

                for p in paquetes_creados:
                    p.factura_id = factura.id
                
                factura.actualizar_total()
                db.session.commit()
                
                registrar_actividad(current_user.id, 'Creó Factura', f'Factura {factura.numero} generada automáticamente')
                flash(f'Paquetes registrados y factura {factura.numero} creada con éxito.', 'success')
                return redirect(url_for('facturas.detalle', id=factura.id))
            
            flash(f'Se registraron {len(paquetes_creados)} paquete(s) con éxito. Costo total estimado: ${costo_total:.2f}', 'success')
            
        return redirect(url_for('paquetes.index'))

    return render_template('paquetes/nuevo.html', clientes=clientes, cliente_id=cliente_id)

@paquetes_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar(id):
    paquete = Paquete.query.get_or_404(id)
    clientes = Cliente.query.filter_by(activo=True).order_by(Cliente.nombre_completo).all()

    if request.method == 'POST':
        peso = int(request.form.get('peso', 0))
        tipo_envio = request.form.get('tipo_envio')
        cliente_id = int(request.form.get('cliente_id'))
        
        cliente = Cliente.query.get(cliente_id)
        tarifa = None
        if cliente and cliente.tarifa_especial:
            if tipo_envio == 'aereo' and cliente.tarifa_especial.aereo is not None:
                tarifa = cliente.tarifa_especial.aereo
            elif tipo_envio == 'maritimo' and cliente.tarifa_especial.maritimo is not None:
                tarifa = cliente.tarifa_especial.maritimo
                
        if tarifa is None:
            tarifa_db = Tarifa.query.filter_by(nombre=tipo_envio).first()
            tarifa = tarifa_db.precio_por_libra if tarifa_db else (6.50 if tipo_envio == 'aereo' else 2.50)

        numero_seguimiento = request.form.get('numero_seguimiento', '').strip()
        if numero_seguimiento and numero_seguimiento != paquete.numero_seguimiento:
            existente = Paquete.query.filter_by(numero_seguimiento=numero_seguimiento).first()
            if existente:
                flash(f'El número de seguimiento "{numero_seguimiento}" ya está registrado en el paquete {existente.tracking_number}.', 'error')
                return redirect(request.url)

        paquete.nombre = request.form.get('nombre').strip()
        paquete.descripcion = request.form.get('descripcion', '').strip()
        paquete.peso = peso
        paquete.tipo_envio = tipo_envio
        paquete.numero_seguimiento = numero_seguimiento
        paquete.estado_rastreo = request.form.get('estado_rastreo', paquete.estado_rastreo)
        paquete.costo = round(peso * tarifa, 2)
        paquete.cliente_id = int(request.form.get('cliente_id'))
        
        if current_user.rol == 'admin':
            registrado_en_str = request.form.get('registrado_en')
            if registrado_en_str:
                try:
                    from datetime import datetime
                    paquete.registrado_en = datetime.strptime(registrado_en_str, '%Y-%m-%dT%H:%M')
                except ValueError:
                    pass

        if paquete.factura:
            paquete.factura.actualizar_total()
            
        db.session.commit()
        
        from models import registrar_actividad
        registrar_actividad(current_user.id, 'Editó Paquete', f'Actualizó paquete {paquete.tracking_number}')
        
        flash('Paquete actualizado.', 'success')
        return redirect(url_for('paquetes.index'))

    return render_template('paquetes/form.html', clientes=clientes, paquete=paquete, cliente_id=paquete.cliente_id)

@paquetes_bp.route('/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar(id):
    if current_user.rol != 'admin':
        flash('Solo los administradores pueden eliminar paquetes.', 'error')
        return redirect(url_for('paquetes.index'))
        
    paquete = Paquete.query.get_or_404(id)
    if paquete.factura_id:
        flash('No se puede eliminar un paquete facturado.', 'error')
        return redirect(url_for('paquetes.index'))
    db.session.delete(paquete)
    db.session.commit()
    
    from models import registrar_actividad
    registrar_actividad(current_user.id, 'Eliminó Paquete', f'Eliminó paquete {paquete.tracking_number}')
    
    flash('Paquete eliminado.', 'success')
    return redirect(url_for('paquetes.index'))

@paquetes_bp.route('/calcular-costo')
@login_required
def calcular_costo():
    peso = int(float(request.args.get('peso', 0)))
    tipo = request.args.get('tipo', 'aereo')
    cliente_id = request.args.get('cliente_id')
    
    tarifa = None
    if cliente_id:
        cliente = Cliente.query.get(int(cliente_id))
        if cliente and cliente.tarifa_especial:
            if tipo == 'aereo' and cliente.tarifa_especial.aereo is not None:
                tarifa = cliente.tarifa_especial.aereo
            elif tipo == 'maritimo' and cliente.tarifa_especial.maritimo is not None:
                tarifa = cliente.tarifa_especial.maritimo
                
    if tarifa is None:
        tarifa_db = Tarifa.query.filter_by(nombre=tipo).first()
        tarifa = tarifa_db.precio_por_libra if tarifa_db else (6.50 if tipo == 'aereo' else 2.50)
        
    costo = round(peso * tarifa, 2)
    return jsonify({'costo': costo, 'tarifa': tarifa})

@paquetes_bp.route('/tarifas-cliente')
@login_required
def tarifas_cliente():
    cliente_id = request.args.get('cliente_id')
    
    t_aereo = Tarifa.query.filter_by(nombre='aereo').first()
    t_maritimo = Tarifa.query.filter_by(nombre='maritimo').first()
    
    precio_aereo = t_aereo.precio_por_libra if t_aereo else 6.50
    precio_maritimo = t_maritimo.precio_por_libra if t_maritimo else 2.50
    
    if cliente_id:
        cliente = Cliente.query.get(int(cliente_id))
        if cliente and cliente.tarifa_especial:
            if cliente.tarifa_especial.aereo is not None:
                precio_aereo = cliente.tarifa_especial.aereo
            if cliente.tarifa_especial.maritimo is not None:
                precio_maritimo = cliente.tarifa_especial.maritimo
                
    return jsonify({
        'aereo': precio_aereo,
        'maritimo': precio_maritimo
    })

@paquetes_bp.route('/<int:id>/historial', methods=['GET', 'POST'])
@login_required
def historial(id):
    paquete = Paquete.query.get_or_404(id)
    if request.method == 'POST':
        estado = request.form.get('estado').strip()
        ubicacion = request.form.get('ubicacion', '').strip()
        comentarios = request.form.get('comentarios', '').strip()

        if estado:
            nuevo_historial = HistorialRastreo(
                paquete_id=paquete.id,
                estado=estado,
                ubicacion=ubicacion,
                comentarios=comentarios,
                creado_por=current_user.id
            )
            paquete.estado_rastreo = estado
            db.session.add(nuevo_historial)
            db.session.commit()
            
            from models import registrar_actividad
            registrar_actividad(current_user.id, 'Actualizó Rastreo', f'Paquete {paquete.tracking_number} a estado "{estado}"')
            
            flash('Historial actualizado correctamente.', 'success')
            return redirect(url_for('paquetes.historial', id=paquete.id))

    return render_template('paquetes/historial.html', paquete=paquete)

@paquetes_bp.route('/exportar')
@login_required
def exportar():
    if current_user.rol != 'admin':
        flash('No tienes permisos para exportar datos.', 'error')
        return redirect(url_for('paquetes.index'))
        
    q = request.args.get('q', '')
    tipo = request.args.get('tipo', '')
    estado = request.args.get('estado', '')
    filtro_fecha = request.args.get('filtro_fecha', '')
    dia = request.args.get('dia', '')
    mes = request.args.get('mes', '')
    semana = request.args.get('semana', '')

    query = Paquete.query.options(joinedload(Paquete.cliente), joinedload(Paquete.factura)).join(Cliente).filter(Cliente.activo == True)
    if q:
        query = query.filter(
            (Paquete.nombre.ilike(f'%{q}%')) |
            (Cliente.nombre_completo.ilike(f'%{q}%')) |
            (Paquete.numero_seguimiento.ilike(f'%{q}%'))
        )
    if tipo:
        query = query.filter(Paquete.tipo_envio == tipo)
    if estado == 'sin_facturar':
        query = query.filter(Paquete.factura_id == None)
    elif estado == 'facturado':
        query = query.filter(Paquete.factura_id != None)

    if filtro_fecha == 'dia' and dia:
        from datetime import datetime, time
        target_date = datetime.strptime(dia, '%Y-%m-%d').date()
        start_date = datetime.combine(target_date, time.min)
        end_date = datetime.combine(target_date, time.max)
        query = query.filter(Paquete.registrado_en >= start_date, Paquete.registrado_en <= end_date)
    elif filtro_fecha == 'mes' and mes:
        year, month = map(int, mes.split('-'))
        import calendar
        from datetime import datetime
        _, last_day = calendar.monthrange(year, month)
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, last_day, 23, 59, 59)
        query = query.filter(Paquete.registrado_en >= start_date, Paquete.registrado_en <= end_date)
    elif filtro_fecha == 'semana' and semana:
        from datetime import datetime, timedelta
        year_str, week_str = semana.split('-W')
        start_date = datetime.strptime(semana + '-1', '%G-W%V-%u')
        end_date = start_date + timedelta(days=7)
        query = query.filter(Paquete.registrado_en >= start_date, Paquete.registrado_en < end_date)

    paquetes = query.order_by(Paquete.registrado_en.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paquetes"

    headers = ['ID', 'Tracking (Sistema)', 'Guía Rastreo', 'Cliente', 'Contenido', 'Peso (lb)', 'Tipo', 'Precio Costo ($)', 'Precio Venta ($)', 'Precio Venta (C$)', 'Ganancia ($)', 'Estado Actual', 'Facturado', 'Fecha Registro']
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='3D5BA0', end_color='3D5BA0', fill_type='solid')

    import re
    def clean_text(val):
        if isinstance(val, str):
            return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
        return val

    for p in paquetes:
        precio_costo = (p.peso * 5.0) if p.tipo_envio == 'aereo' else (p.peso * 1.6)
        precio_venta = p.costo or 0
        ganancia = precio_venta - precio_costo
        
        row = [
            p.id,
            p.tracking_number,
            p.numero_seguimiento or '',
            p.cliente.nombre_completo,
            p.nombre,
            p.peso,
            p.tipo_envio.upper() if p.tipo_envio else '',
            precio_costo,
            precio_venta,
            precio_venta * 37,
            ganancia,
            p.estado_rastreo.replace('_', ' ').title() if p.estado_rastreo else '',
            'SÍ' if p.factura_id else 'NO',
            p.registrado_en.strftime('%Y-%m-%d %H:%M') if p.registrado_en else ''
        ]
        ws.append([clean_text(v) for v in row])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = (max_length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from models import get_local_now
    filename = f"Paquetes_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

import urllib.request
import urllib.parse
import http.cookiejar
import time
import unicodedata
import json

class AereomarClient:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.cj = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(self.cj))
        self.logged_in = False
        self.last_login_time = 0
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }

    def login(self):
        try:
            home_url = 'https://aereomarexpress.multitrack.trackingpremium.us/'
            login_url = 'https://aereomarexpress.multitrack.trackingpremium.us/login_check'
            
            # GET inicial
            req_home = urllib.request.Request(home_url, headers=self.headers)
            self.opener.open(req_home, timeout=10)
            
            # POST credenciales
            login_data = urllib.parse.urlencode({
                '_username': self.username,
                '_password': self.password
            }).encode('utf-8')
            
            headers_post = dict(self.headers)
            headers_post['Content-Type'] = 'application/x-www-form-urlencoded'
            headers_post['Referer'] = home_url
            
            req_login = urllib.request.Request(login_url, data=login_data, headers=headers_post)
            self.opener.open(req_login, timeout=15)
            self.logged_in = True
            self.last_login_time = time.time()
            return True
        except Exception as e:
            print("Error en login Aereomar:", e)
            self.logged_in = False
            return False

    def ensure_logged_in(self):
        if not self.logged_in or (time.time() - self.last_login_time > 1500):
            return self.login()
        return True

    def fetch_whrec_trackings(self, whrec_ruta):
        if not self.ensure_logged_in():
            return []
        
        try:
            url = f"https://aereomarexpress.multitrack.trackingpremium.us{whrec_ruta}"
            req = urllib.request.Request(url, headers=self.headers)
            html = self.opener.open(req, timeout=10).read().decode('utf-8', errors='ignore')
            
            if 'login_check' in html or len(html) < 15000:
                self.login()
                req = urllib.request.Request(url, headers=self.headers)
                html = self.opener.open(req, timeout=10).read().decode('utf-8', errors='ignore')
                
            rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL)
            results = []
            for r in rows:
                clean_cells = re.findall(r'<td[^>]*>(.*?)</td>', r, re.DOTALL)
                cells = [re.sub(r'<[^>]+>', '', c).strip() for c in clean_cells]
                if len(cells) >= 8 and 'XP' in cells[0]:
                    pkg_code = cells[0]
                    desc = cells[5] if len(cells) > 5 else ""
                    carrier_tracking = cells[7] if len(cells) > 7 else ""
                    raw_weight = cells[11] if len(cells) > 11 else ""
                    
                    weight = 0
                    w_m = re.search(r'([\d\.]+)', raw_weight)
                    if w_m:
                        try:
                            weight = float(w_m.group(1))
                        except Exception:
                            weight = 0
                            
                    results.append({
                        "pkg_code": pkg_code,
                        "carrier_tracking": carrier_tracking,
                        "description": desc,
                        "weight": weight
                    })
            return results
        except Exception as e:
            print("Error extrayendo whrec trackings:", e)
            return []

aereomar_client = AereomarClient(
    os.environ.get('AEREOMAR_USER', 'enviosbjjexpress@gmail.com'),
    os.environ.get('AEREOMAR_PASS', 'Julio100!')
)

@paquetes_bp.route('/fetch_aereomar', methods=['POST'])
@login_required
def fetch_aereomar():
    data = request.get_json() or {}
    tracking_full = data.get('tracking_number', '').strip()
    search_type_hint = data.get('search_type', '').strip().lower() # 'warehouse', 'tracking', or ''
    
    if not tracking_full:
        return jsonify({"success": False, "error": "No se proporcionó número para buscar"})
    
    def normalize_str(s):
        if not s:
            return ""
        s = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
        return re.sub(r'[^a-zA-Z0-9\s]', ' ', s).lower().strip()
    
    base_tracking_match = re.match(r'^(WR\d+)', tracking_full, re.IGNORECASE)
    is_warehouse_query = (search_type_hint == 'warehouse') or bool(base_tracking_match) or tracking_full.upper().startswith('WR')
    
    def fetch_data(search_type, search_number):
        url = f'https://aereomarexpress.multitrack.trackingpremium.us/tracking/search?type={search_type}&number={search_number}&user=0&recibo=0&guia=0&consol=0'
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)', 'X-Requested-With': 'XMLHttpRequest'})
        resp = urllib.request.urlopen(req, timeout=10).read().decode('utf-8')
        return json.loads(resp)

    try:
        json_data = None
        # Si es tipo warehouse (WR), intentamos type=2 primero
        if is_warehouse_query:
            wr_code = base_tracking_match.group(1).upper() if base_tracking_match else tracking_full.upper()
            try:
                json_data = fetch_data(2, wr_code)
            except Exception:
                json_data = None
            if not (json_data and len(json_data) >= 6 and json_data[5]):
                try:
                    json_data = fetch_data(4, tracking_full) # fallback tracking
                except Exception:
                    pass
        else:
            # Si es tracking normal, intentamos type=4 primero
            try:
                json_data = fetch_data(4, tracking_full)
            except Exception:
                json_data = None
            if not (json_data and len(json_data) >= 6 and json_data[5]):
                try:
                    json_data = fetch_data(2, tracking_full) # fallback warehouse
                except Exception:
                    pass
            if not (json_data and len(json_data) >= 6 and json_data[5]):
                try:
                    json_data = fetch_data(3, tracking_full) # fallback recibo/paquete
                except Exception:
                    pass

        if json_data and len(json_data) >= 6 and json_data[5]:
            basic_data = json_data[5]
            raw_desc = basic_data[0].strip() if len(basic_data) > 0 and basic_data[0] else ""
            weight = basic_data[1] if len(basic_data) > 1 else 0
            
            ship_type_str = str(basic_data[4]).lower() if len(basic_data) > 4 else ""
            ship_type = "aereo"
            if "mar" in ship_type_str or "maritimo" in ship_type_str:
                ship_type = "maritimo"
                
            # Extraer Warehouse
            warehouse = ""
            whrec_ruta = ""
            if len(json_data) > 1 and json_data[1] and isinstance(json_data[1], list):
                if len(json_data[1]) > 0 and 'number' in json_data[1][0]:
                    warehouse = str(json_data[1][0]['number']).strip()
                if len(json_data[1]) > 0 and 'ruta' in json_data[1][0]:
                    whrec_ruta = str(json_data[1][0]['ruta']).strip()
                    
            if not warehouse and is_warehouse_query:
                warehouse = base_tracking_match.group(1).upper() if base_tracking_match else tracking_full.upper()

            # Consultar tracking(s) original(es) de la tienda usando la sesión autenticada
            whrec_items = []
            if whrec_ruta:
                try:
                    whrec_items = aereomar_client.fetch_whrec_trackings(whrec_ruta)
                except Exception as e_wh:
                    print("Error consultando whrec autenticado:", e_wh)

            # Extraer lista de paquetes contenidos en este recibo / warehouse
            packages_list = []
            primary_tracking = ""

            if whrec_items:
                # Usar los datos autenticados con trackings reales de la tienda
                for item in whrec_items:
                    c_track = item.get('carrier_tracking') or item.get('pkg_code')
                    packages_list.append({
                        "numero_seguimiento": c_track,
                        "pkg_code": item.get('pkg_code'),
                        "descripcion": item.get('description') or raw_desc,
                        "peso": item.get('weight') or weight
                    })
                if packages_list:
                    primary_tracking = packages_list[0]['numero_seguimiento']
            else:
                # Fallback al json público
                if len(json_data) > 3 and json_data[3] and isinstance(json_data[3], list):
                    for p_item in json_data[3]:
                        if isinstance(p_item, dict):
                            p_num = str(p_item.get('number', '')).strip()
                            p_desc = str(p_item.get('description', '')).strip()
                            p_weight = p_item.get('weight', 0)
                            if p_num:
                                packages_list.append({
                                    "numero_seguimiento": p_num,
                                    "descripcion": p_desc,
                                    "peso": p_weight
                                })

                if packages_list:
                    primary_tracking = packages_list[0]['numero_seguimiento']
                elif not is_warehouse_query:
                    primary_tracking = tracking_full

            # Si el usuario buscó por tracking original, respetamos ese tracking exacto
            if not is_warehouse_query and tracking_full:
                primary_tracking = tracking_full

            # Parsear descripción: separar alias del cliente y producto
            nombre_limpio = raw_desc
            alias_cliente = ""
            if "//" in raw_desc:
                partes = raw_desc.split("//", 1)
                alias_cliente = partes[0].strip()
                nombre_limpio = partes[1].strip()
            elif " - " in raw_desc:
                partes = raw_desc.split(" - ", 1)
                alias_cliente = partes[0].strip()
                nombre_limpio = partes[1].strip()
            elif "/" in raw_desc:
                partes = raw_desc.split("/", 1)
                alias_cliente = partes[0].strip()
                nombre_limpio = partes[1].strip()
            else:
                alias_cliente = raw_desc

            # Smart Client Match en base de datos local
            matched_cliente = None
            if alias_cliente:
                norm_alias = normalize_str(alias_cliente)
                stop_words = {'amazon', 'shein', 'temu', 'ebay', 'walmart', 'bjj', 'express', 'box', 'casillero', 'paquete', 'usa', 'miami'}
                alias_words = [w for w in norm_alias.split() if len(w) > 2 and w not in stop_words]
                
                clientes_activos = Cliente.query.filter_by(activo=True).all()
                best_score = 0
                
                for c in clientes_activos:
                    c_norm = normalize_str(c.nombre_completo)
                    c_words = [w for w in c_norm.split() if len(w) > 2]
                    
                    if c_norm and (c_norm in norm_alias or norm_alias in c_norm):
                        score = 100
                    else:
                        shared = set(alias_words).intersection(set(c_words))
                        if len(shared) >= 2:
                            score = 80 + len(shared) * 5
                        elif len(shared) == 1 and len(c_words) <= 2:
                            score = 60
                        elif len(shared) == 1:
                            score = 40
                        else:
                            score = 0
                            
                    if score > best_score and score >= 40:
                        best_score = score
                        matched_cliente = c

            cliente_data = None
            if matched_cliente:
                tarifa_esp = getattr(matched_cliente, 'tarifa_especial', None)
                cliente_data = {
                    "id": matched_cliente.id,
                    "nombre": matched_cliente.nombre_completo,
                    "cedula": matched_cliente.cedula,
                    "tarifa_aereo": tarifa_esp.aereo if tarifa_esp and tarifa_esp.aereo else None,
                    "tarifa_maritimo": tarifa_esp.maritimo if tarifa_esp and tarifa_esp.maritimo else None
                }

            return jsonify({
                "success": True,
                "warehouse": warehouse,
                "numero_seguimiento": primary_tracking,
                "nombre": raw_desc or nombre_limpio,
                "descripcion": raw_desc or nombre_limpio,
                "descripcion_completa": raw_desc,
                "peso": weight,
                "tipo_envio": ship_type,
                "cliente": cliente_data,
                "paquetes": packages_list
            })
        else:
            return jsonify({"success": False, "error": "No se encontraron datos en AereoMar para este número"})
    except Exception as e:
        return jsonify({"success": False, "error": f"Error consultando AereoMar: {str(e)}"})


